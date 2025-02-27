import imaplib
import base64
import os
from dotenv import load_dotenv
import time
from configparser import ConfigParser, ExtendedInterpolation
import wget
import urllib
import xlrd
import uuid
import csv
from bson.objectid import ObjectId
from datetime import datetime
import json
from datetime import datetime
import requests
from difflib import get_close_matches
from requests import post, get, delete
import sys
import time
import xlwt
import xlutils
import subprocess
from xlutils.copy import copy
import shutil
import re
from xlrd import open_workbook
from xlutils.copy import copy as xl_copy
import logging
import logging.handlers
import time
from logging.handlers import TimedRotatingFileHandler
import xlsxwriter
import argparse
import sys
from os import path
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import gdown
from mimetypes import guess_extension
from datetime import datetime, timedelta
import datetime as dt
import pytz


load_dotenv()

# get current working directory
currentDirectory = os.getcwd()

# Read config file 
config = ConfigParser()
config.read('common_config/config_mentoring.ini')


# email regex
regex = "\"?([-a-zA-Z0-9.`?{}]+@\w+\.\w+)\"?"

# Global variable declaration
criteriaLookUp = dict()
millisecond = None
programNameInp = None
environment = None
observationId = None
solutionName = None
pointBasedValue = None
entityType = None
allow_multiple_submissions = None
scopeEntityType = ""
programName = None
userEntity = None
roles = ""
mainRole = ""
dictCritLookUp = {}
isProgramnamePresent = None
solutionLanguage = None
keyWords = None
entityTypeId = None
solutionDescription = None
creator = None
data = None
projectServiceLoginId = None
criteriaName = None
solutionId = None
API_log = None
listOfFoundRoles = []
entityToUpload = None
programID = None
programExternalId = None
programDescription = None
criteriaLookUp = dict()
themesSheetList = []
themeRubricFileObj = dict()
criteriaLevelsReport = False
ecm_sections = dict()
criteriaLevelsCount = 0
numberOfResponses = 0
criteriaIdNameDict = dict()
criteriaLevels = list()
matchedShikshalokamLoginId = None
scopeEntities = []
scopeRoles = []
countImps = 0
ecmToSection = dict()
entitiesPGM = []
stateEntitiesPGM = []
entitiesPGMID = []
entitiesType = []
solutionRolesArr = []
startDateOfResource = None
endDateOfResource = None
startDateOfProgram = None
endDateOfProgram = None
rolesPGM =None
mainRole = None
solutionRolesArray = []
solutionStartDate = ""
solutionEndDate = ""
projectCreator = ""
orgIds = []
OrgName = []
ccRootOrgName = None
ccRootOrgId  = None
certificatetemplateid = None
question_sequence_arr = []


# Generate access token for the APIs. 
def generateAccessToken():
    # production search user api - start
    headerKeyClockUser = {'Content-Type': config.get(environment, 'content-type')}
    loginBody = {
        'email' : os.getenv('MentoringEmail' + environment ),
        'password' :os.getenv('MentoringPassword' + environment )
    }
    print(loginBody)
    responseKeyClockUser = requests.post(config.get(environment, 'mentoringuserhost') + config.get(environment, 'userlogin'), headers=headerKeyClockUser, json=loginBody)  
    messageArr = []
    messageArr.append("URL : " + str(config.get(environment, 'userlogin')))
    messageArr.append("Body : " + str(os.getenv('MentoringkeyclockapibodyDEV')))
    messageArr.append("Status Code : " + str(responseKeyClockUser.status_code))
    if responseKeyClockUser.status_code == 200:
        responseKeyClockUser = responseKeyClockUser.json()
        accessTokenUser = responseKeyClockUser['result']['access_token']
        print("--->Access Token Generated!")
        return accessTokenUser
    
    print("Error in generating Access token")
    print("Status code : " + str(responseKeyClockUser.status_code))
    terminatingMessage("Please check API logs.")


def AdminCreate(UserFile):
    header_create_user = {
            'Content-type':'application/json',
            'internal_access_token':os.getenv("MentoringinternalAccessToken" + environment)
        }
    AdminSheet = xlrd.open_workbook(UserFile,on_demand=True)
    AdminData = AdminSheet.sheet_names()
    for sheetEnv in AdminData:
       if sheetEnv == 'Admin Role':
           detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
           keysEnv = [detailsEnvSheet.cell(0, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
           for row_index_env in range(1, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)}
                email = dictDetailsEnv['email']
                password = dictDetailsEnv['password']
                name = dictDetailsEnv['name']
                secret_code = os.getenv("secret_code" + environment)
                user_payload = {
                        "email": email,
                        "password": password,
                        "name": name,
                        "secret_code":secret_code,                       
                }
                url_create_user_api = config.get(environment, 'mentoringuserhost') + config.get(environment, 'AdminUserCreate')
                response = requests.post(
                        url=url_create_user_api,
                        headers=header_create_user,
                        data=json.dumps(user_payload)
                    )
                if response.status_code == 201:
                    print(f"User Admin creation succeeded for title: {name}")
                else:
                    print(f"User Admin creation failed for title: {name}")
                    print(f"Response: {response.text}")
                    break


def UserCreate(UserFile):
    header_create_user = {
            'Content-type':'application/json',
            'internal_access_token':'internal_access_token'
        }
    AdminSheet = xlrd.open_workbook(UserFile,on_demand=True)
    AdminData = AdminSheet.sheet_names()
    for sheetEnv in AdminData:
       if sheetEnv == 'User Role':
           detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
           keysEnv = [detailsEnvSheet.cell(0, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
           for row_index_env in range(1, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)}
                # emailValidationOTP = getOTP(dictDetailsEnv)
                email = dictDetailsEnv['Email']
                password = dictDetailsEnv['Password']
                name = dictDetailsEnv['Name']
                role = dictDetailsEnv['Role']
                user_payload = {
                        "email": email,
                        "password": password,
                        "name": name,
                        "role":role,
                        # "otp":emailValidationOTP
                       
                }
                url_create_user_api = config.get(environment, 'mentoringuserhost') + config.get(environment, 'UserCreate')
                response = requests.post(
                        url=url_create_user_api,
                        headers=header_create_user,
                        data=json.dumps(user_payload)
                    )
                if response.status_code == 201:
                    print(f"User creation succeeded for title: {name}")
                else:
                    print(f"User failed for title: {name}")
                    print(f"Response: {response.text}")
                    break
                

def createOrginzation(UserFile,accessToken):
    header_create_user = {
            'Content-type':'application/json',
            'X-auth-token': 'bearer ' + accessToken,
            'internal_access_token':'internal_access_token'
        }
    AdminSheet = xlrd.open_workbook(UserFile,on_demand=True)
    AdminData = AdminSheet.sheet_names()
    for sheetEnv in AdminData:
       if sheetEnv == 'Organization':
           detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
           keysEnv = [detailsEnvSheet.cell(0, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
           for row_index_env in range(1, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)}
                name = dictDetailsEnv['Name']
                code = dictDetailsEnv['Code']
                description = dictDetailsEnv['Description']
                domains = dictDetailsEnv['Domains']
                adminRole = dictDetailsEnv['Admin Email']
                user_payload = {
                        "name": name,
                        "code": code,
                        "description": description,
                        "domains":domains,
                        "admin_email":adminRole
                }
                url_create_user_api = config.get(environment, 'mentoringuserhost') + config.get(environment, 'OrganizationCreate')
                response = requests.post(
                        url=url_create_user_api,
                        headers=header_create_user,
                        data=json.dumps(user_payload)
                    )
                if response.status_code == 201:
                    print(f"Organization creation succeeded for name: {name}")
                else:
                    print(f"Organization creation failed for name: {name}")
                    print(f"Response: {response.text}")
                    break


def setOrgPolicies(UserFile,UserRole):
    AdminSheet = xlrd.open_workbook(UserFile,on_demand=True)
    AdminData = AdminSheet.sheet_names()
    for sheetEnv in AdminData:
       if sheetEnv == 'Policies':
           detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
           keysEnv = [detailsEnvSheet.cell(0, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)]
           for row_index_env in range(1, detailsEnvSheet.nrows):
                dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value for col_index_env in range(detailsEnvSheet.ncols)}
                sessionVisibilityPolicy = dictDetailsEnv['Session Visibility Policy']
                mentorVisibilityPolicy = dictDetailsEnv['Mentor Visibility Policy']
                externalSessionVisibilityPolicy = dictDetailsEnv['External Session Visibility Policy']
                externalMentorVisibilityPolicy = dictDetailsEnv['External Mentor Visibility Policy']
                allowMentorOverride = dictDetailsEnv['Allow Mentor Override']
                UserRoleId = int(dictDetailsEnv['User Mapping ID'])
                UserRoleDetails = UserRole[(UserRoleId - 1)]
                accessTokenSession = GenerateAccessTokenSession(UserRoleDetails)

                Policiespayload = {
                         "session_visibility_policy": sessionVisibilityPolicy,
                         "mentor_visibility_policy": mentorVisibilityPolicy,
                         "external_session_visibility_policy": externalSessionVisibilityPolicy,
                         "external_mentor_visibility_policy": externalMentorVisibilityPolicy,
                         "allow_mentor_override": allowMentorOverride
                }
                header_create_user = {
                'Content-type':'application/json',
                'X-auth-token': 'bearer ' + accessTokenSession,
                }
                url_create_user_api = config.get(environment, 'mentoringuserhost') + config.get(environment, 'OrgPolicieCreate')
                response = requests.post(
                        url=url_create_user_api,
                        headers=header_create_user,
                        data=json.dumps(Policiespayload)
                    )
                if response.status_code == 200:
                    print(f"Organization creation succeeded")
                else:
                    print(f"Organization creation failed")
                    print(f"Response: {response.text}")
                    break


def GenerateAccessTokenSession(UserFile):
    # production search user api - start
    headerKeyClockUser = {'Content-Type': config.get(environment, 'content-type')}
    loginBody = {
        'email' : UserFile['Email'],
        'password' : UserFile['Password']
    }
    responseKeyClockUser = requests.post(config.get(environment, 'mentoringuserhost') + config.get(environment, 'userlogin'), headers=headerKeyClockUser, json=loginBody)  
    if responseKeyClockUser.status_code == 200:
        responseKeyClockUser = responseKeyClockUser.json()
        accessTokenUser = responseKeyClockUser['result']['access_token']
        print("--->Access Token Generated!")
        return accessTokenUser
    
    print("Error in generating Access token!")
    print("Status code : " + str(responseKeyClockUser.status_code))
    terminatingMessage("Please check API logs.")

def sessionsFlow(UserFile, UserRole):
    AdminSheet = xlrd.open_workbook(UserFile, on_demand=True)
    AdminData = AdminSheet.sheet_names()

    for sheetEnv in AdminData:
        if sheetEnv == 'Sessions':
            detailsSessionSheet = AdminSheet.sheet_by_name(sheetEnv)
            keysEnvSessions = [detailsSessionSheet.cell(0, col_index_env).value for col_index_env in range(detailsSessionSheet.ncols)]
            
            for row_index_env in range(1, detailsSessionSheet.nrows):
                dictSessionEnv = {
                    keysEnvSessions[col_index_env]: detailsSessionSheet.cell(row_index_env, col_index_env).value 
                    for col_index_env in range(detailsSessionSheet.ncols)
                }
                
                # Handling Date Column (Assuming Date is in a column named "Date")
                if "Date" in dictSessionEnv:
                    col_index_date = keysEnvSessions.index("Date")
                    cell_value = detailsSessionSheet.cell(row_index_env, col_index_date).value
                    cell_type = detailsSessionSheet.cell_type(row_index_env, col_index_date)

                    if cell_type == xlrd.XL_CELL_DATE:
                        dictSessionEnv["Date"] = xlrd.xldate_as_datetime(cell_value, AdminSheet.datemode).strftime("%d-%m-%Y")
                action = dictSessionEnv["Action"].strip().lower()
                if action in ["create", "edit", "delete"]:
                    UserId = int(dictSessionEnv["User Mapping ID"])
                    UserRoleDetails = UserRole[UserId - 1]
                    accessTokenSession = GenerateAccessTokenSession(UserRoleDetails)

                    if action == "create":
                        createSession(dictSessionEnv, accessTokenSession)
                    elif action == "editt":
                        EditSession(dictSessionEnv, accessTokenSession)
                    elif action == "deletet":
                        deleteSession(dictSessionEnv, accessTokenSession)
                else:
                    terminatingMessage("No Valid Action Found")



def CreatingReport(UserFile, UserRole):
    AdminSheet = xlrd.open_workbook(UserFile, on_demand=True)
    AdminData = AdminSheet.sheet_names()

    for sheetEnv in AdminData:
        if sheetEnv == 'Sessions Report':
            detailsSessionSheet = AdminSheet.sheet_by_name(sheetEnv)
            keysEnvSessions = [detailsSessionSheet.cell(0, col_index_env).value for col_index_env in range(detailsSessionSheet.ncols)]
            
            for row_index_env in range(1, detailsSessionSheet.nrows):
                dictSessionEnv = {
                    keysEnvSessions[col_index_env]: detailsSessionSheet.cell(row_index_env, col_index_env).value 
                    for col_index_env in range(detailsSessionSheet.ncols)
                }
                
                action = dictSessionEnv["Action"].strip().lower()
                if action in ["create", "edit", "delete"]:
                    UserId = int(dictSessionEnv["User Mapping ID"])
                    UserRoleDetails = UserRole[UserId - 1]
                    accessTokenSession = GenerateAccessTokenSession(UserRoleDetails)

                    if action == "create":
                        SessionIdToStart = createSessionForReport(dictSessionEnv, accessTokenSession)
                        startSessions(SessionIdToStart,accessTokenSession )
                    elif action == "edit":
                        EditSessionForReport(dictSessionEnv, accessTokenSession)
                    elif action == "delete":
                        deleteSessionForReport(dictSessionEnv, accessTokenSession)
                else:
                    terminatingMessage("No Valid Action Found")

def startSessions(SessionIdToStart,accessTokenSession):
    header_session_start = {
        'X-auth-token': 'bearer ' + accessTokenSession,
        'Content-type':'application/json'
    }

    sessionIDInString = str(SessionIdToStart)
    url_start_session = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionStart') + "/" + sessionIDInString
    response = requests.post(
                url=url_start_session,
                headers=header_session_start,
                )
    if response.status_code == 200:
        print(f"Session Started Successfully  for ID: {SessionIdToStart}")
    else:
        print(f"Session Failed Start for ID: {SessionIdToStart}")
        print(f"Response: {response.text}")


def deleteSessionForReport(row,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }

    id = row["id"]
    strId = str(id)
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionCreate') + "/" + strId
    response = requests.delete(
                url=url_create_user_api,
                headers=header_create_user,
                )
    if response.status_code == 201:
        print(f"Session Deletion succeeded for ID: {id}")
    else:
        print(f"Session Deletion failed for ID: {id}")
        print(f"Response: {response.text}")


def createSessionForReport(row,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }

    current = dt.datetime.now()

    DateForEpoch = f"{current.day}-{current.month}-{current.year}"

    TimeForEpoch = f"{current.hour}:{current.minute} Hrs"

    Action = row['Action']
    id = row['id']
    title = row['title']
    description = row['description']
    type = row['type']
    mentorEmail = row['Mentor Email']
    MentorID = FetchMentorId(mentorEmail,accessToken)
    MenteesEmail = row['Mentees Email']
    MenteeID = FetchMenteeId(MenteesEmail,accessToken)
    dates = convert_to_epoch(DateForEpoch,TimeForEpoch,row['Time Zone'],row['Duration'])
    time_zone = row['Time Zone']
    recommended_for = row['recommended_for']
    categories = row['categories']
    medium = row['medium']
    platform = row['Meeting Platform']
    link = row['Meeting Link']
    passCode = row['Meeting Passcode']
    strMent = int(MenteeID)
    strMentor = int(MentorID)
    session_payload = {
        "Action": Action,
        "id": id,
        "title": title,
        "description": description,
        "mentees": [str(strMent)],
        "mentor_id":str(strMentor),
        "type": type,
        "start_date": str(dates[0]),
        "end_date": str(dates[1]),
        "recommended_for": [recommended_for],
        "categories": [categories],
        "medium": [medium],
        "time_zone": time_zone,
        "meeting_info": {
            "platform":platform,
            "link":link,
            "value":passCode,
            "meta":{}
            },
        }
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionCreate')
    response = requests.post(
                    url=url_create_user_api,
                    headers=header_create_user,
                    data=json.dumps(session_payload)
    )
    if response.status_code == 201:
        responseJson = response.json()
        SessionID = responseJson["result"]["id"]
        print(f"Session creation For Report succeeded for title: {title}")
        return SessionID
    else:
        print(f"Session creation For Report failed for title: {title}")
        print(f"Response: {response.text}")


def EditSessionForReport(row,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }


    current = dt.datetime.now()

    DateForEpoch = f"{current.day}-{current.month}-{current.year}"

    TimeForEpoch = f"{current.hour}:{current.minute} Hrs"
    
    id = row['id']       
    title = row['title']
    description = row['description']
    type = row['type']
    mentorEmail = row['Mentor Email']
    MentorID = FetchMentorId(mentorEmail,accessToken)
    MenteesEmail = row['Mentees Email']
    MenteeID = FetchMenteeId(MenteesEmail,accessToken)
    dates = convert_to_epoch(DateForEpoch,TimeForEpoch,row['Time Zone'],row['Duration'])
    time_zone = row['Time Zone']
    recommended_for = row['recommended_for']
    categories = row['categories']
    medium = row['medium']
    platform = row['Meeting Platform']
    link = row['Meeting Link']
    passCode = row['Meeting Passcode']
    strMent = int(MenteeID)
    strMentor = int(MentorID)
    intID = str(id)
    session_payload = {
        "id": intID,
        "title": title,
        "description": description,
        "mentees": [str(strMent)],
        "mentor_id":str(strMentor),
        "type": type,
        "start_date": str(dates[0]),
        "end_date": str(dates[1]),
        "recommended_for": [recommended_for],
        "categories": [categories],
        "medium": [medium],
        "time_zone": time_zone,
        "meeting_info": {
            "platform":platform,
            "link":link,
            "value":passCode,
            "meta":{}
            },
        }
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionCreate') + "/" + intID
    response = requests.post(
                url=url_create_user_api,
                headers=header_create_user,
                data=json.dumps(session_payload)
                )

    if response.status_code == 201:
        print(f"Session Edited For Report succeeded for title: {title}")
    else:
        print(f"Session Edited For Report failed for title: {title}")
        print(f"Response: {response.text}")

def sessionTokenGeneration(UserFile):
    AdminSheet = xlrd.open_workbook(UserFile,on_demand=True)
    AdminData = AdminSheet.sheet_names()
    for sheetEnv in AdminData:
        detailsEnvUserSheet = wbPgm.sheet_by_name("User Role")
        keysEnvUserRole = [detailsEnvUserSheet.cell(0, col_index_env).value for col_index_env in range(detailsEnvUserSheet.ncols)]
        data = []
        for row_index_env in range(1, detailsEnvUserSheet.nrows):
            dictUserRoleEnv = {keysEnvUserRole[col_index_env]: detailsEnvUserSheet.cell(row_index_env, col_index_env).value for col_index_env in range(detailsEnvUserSheet.ncols)}            
            data.append(dictUserRoleEnv)
    return data

def terminatingMessage(message):
    print(message)
    exit(1)
  
def convert_to_epoch(date, time, tz, duration):
    setTime = []
    dt_str = f"{date} {time[:-4]}"  # Removing ' Hrs'
    dt_format = "%d-%m-%Y %H:%M"
    dt = datetime.strptime(dt_str, dt_format)
        
    timezone = pytz.timezone(tz)
    dt = timezone.localize(dt)
        
        # Convert to epoch
    start_epoch = int(dt.timestamp())
        
        # Calculate end time
    end_dt = dt + timedelta(minutes=duration)
    end_epoch = int(end_dt.timestamp())    
    setTime.append(start_epoch)
    setTime.append(end_epoch)
    return setTime


def deleteSession(row,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }

    id = row["id"]
    strId = str(id)
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionCreate') + "/" + strId
    response = requests.delete(
                url=url_create_user_api,
                headers=header_create_user,
                )
    if response.status_code == 201:
        print(f"Session Deletion succeeded for ID: {id}")
    else:
        print(f"Session Deletion failed for ID: {id}")
        print(f"Response: {response.text}")


# Function to upload evidence to cloud
def FetchMentorId(MentorEMail, accessToken):
    mentorListUrlApi = config.get(environment, 'mentoringthost') + config.get(environment, 'MentorList')
    headermentorLisApi = {
        'Content-Type': config.get(environment, 'Content-Type'),
        'X-auth-token': 'bearer ' + accessToken,
    }

    b = base64.b64encode(bytes(MentorEMail, 'utf-8'))
    k = str(b)
    h = k.split("'")

    responseMentorLisApiApi = requests.get(url=mentorListUrlApi, headers=headermentorLisApi, params={"search": h[1]})
    if responseMentorLisApiApi.status_code == 200:
            responseMentorLisApiApiJSON = responseMentorLisApiApi.json()
            id = responseMentorLisApiApiJSON["result"]["data"][0]["id"]
            return id
    else:
        terminatingMessage("Mentor List URL Failed")


# Function to upload evidence to cloud
def FetchMenteeId(MenteeEMail, accessToken):
    mentorListUrlApi = config.get(environment, 'mentoringthost') + config.get(environment, 'MenteeList')
    headermentorLisApi = {
        'Content-Type': config.get(environment, 'Content-Type'),
        'X-auth-token': 'bearer ' + accessToken,
    }

    b = base64.b64encode(bytes(MenteeEMail, 'utf-8'))
    k = str(b)
    h = k.split("'")

    responseMentorLisApiApi = requests.get(url=mentorListUrlApi, headers=headermentorLisApi, params={"search": h[1]})
    if responseMentorLisApiApi.status_code == 200:
            responseMentorLisApiApiJSON = responseMentorLisApiApi.json()
            id = responseMentorLisApiApiJSON["result"]["data"][0]["id"]
            return id
    else:
        terminatingMessage("Mentee List URL Failed")

def createSession(row,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }
    Action = row['Action']
    id = row['id']
    title = row['title']
    description = row['description']
    type = row['type']
    mentorEmail = row['Mentor Email']
    MentorID = FetchMentorId(mentorEmail,accessToken)
    MenteesEmail = row['Mentees Email']
    MenteeID = FetchMenteeId(MenteesEmail,accessToken)
    dates = convert_to_epoch(row['Date'],row['Time'],row['Time Zone'],row['Duration'])
    time_zone = row['Time Zone']
    recommended_for = row['recommended_for']
    categories = row['categories']
    medium = row['medium']
    platform = row['Meeting Platform']
    link = row['Meeting Link']
    passCode = row['Meeting Passcode']
    strMent = int(MenteeID)
    strMentor = int(MentorID)
    session_payload = {
        "Action": Action,
        "id": id,
        "title": title,
        "description": description,
        "mentees": [str(strMent)],
        "mentor_id":str(strMentor),
        "type": type,
        "start_date": str(dates[0]),
        "end_date": str(dates[1]),
        "recommended_for": [recommended_for],
        "categories": [categories],
        "medium": [medium],
        "time_zone": time_zone,
        "meeting_info": {
            "platform":platform,
            "link":link,
            "value":passCode,
            "meta":{}
            },
        }
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionCreate')
    response = requests.post(
                    url=url_create_user_api,
                    headers=header_create_user,
                    data=json.dumps(session_payload)
    )
    if response.status_code == 201:
        print(f"Session creation succeeded for title: {title}")
    else:
        print(f"Session creation failed for title: {title}")
        print(f"Response: {response.text}")


def EditSession(row,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }
    
    id = row['id']       
    title = row['title']
    description = row['description']
    type = row['type']
    mentorEmail = row['Mentor Email']
    MentorID = FetchMentorId(mentorEmail,accessToken)
    MenteesEmail = row['Mentees Email']
    MenteeID = FetchMenteeId(MenteesEmail,accessToken)
    dates = convert_to_epoch(row['Date'],row['Time'],row['Time Zone'],row['Duration'])
    time_zone = row['Time Zone']
    recommended_for = row['recommended_for']
    categories = row['categories']
    medium = row['medium']
    platform = row['Meeting Platform']
    link = row['Meeting Link']
    passCode = row['Meeting Passcode']
    strMent = int(MenteeID)
    strMentor = int(MentorID)
    intID = str(id)
    session_payload = {
        "id": intID,
        "title": title,
        "description": description,
        "mentees": [str(strMent)],
        "mentor_id":str(strMentor),
        "type": type,
        "start_date": str(dates[0]),
        "end_date": str(dates[1]),
        "recommended_for": [recommended_for],
        "categories": [categories],
        "medium": [medium],
        "time_zone": time_zone,
        "meeting_info": {
            "platform":platform,
            "link":link,
            "value":passCode,
            "meta":{}
            },
        }
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'sessionCreate') + "/" + intID
    response = requests.post(
                url=url_create_user_api,
                headers=header_create_user,
                data=json.dumps(session_payload)
                )

    if response.status_code == 201:
        print(f"Session creation succeeded for title: {title}")
    else:
        print(f"Session creation failed for title: {title}")
        print(f"Response: {response.text}")

# Function to upload evidence to cloud
def getPreSignedUrl(accessToken):
    preSignedUrlApi = config.get(environment, 'mentoringthost') + config.get(environment, 'fetchPreSignedUrl')
    headerPreSignedUrlApi = {
        'Content-Type': config.get(environment, 'Content-Type'),
        'X-auth-token': 'bearer ' + accessToken,
    }
   
    Folder_Path = os.listdir('MentoringDataCSV/')
    if "Update Roles.csv" in Folder_Path:
        QueryParams = "Update Roles.csv"
    else:
        terminatingMessage("Update Roles.csv  File is not")
    responsegetPreSignedUrlApi = requests.get(url=preSignedUrlApi, headers=headerPreSignedUrlApi, params={"fileName": QueryParams})
    if responsegetPreSignedUrlApi.status_code == 200:
            responsegetPreSignedUrlApi = responsegetPreSignedUrlApi.json()
            files = responsegetPreSignedUrlApi["result"]["signedUrl"]
    else:
        terminatingMessage("PreSigned URL Failed")

    csv_file_path = './MentoringDataCSV/Update Roles.csv'
    with open(csv_file_path, 'rb') as binary_file:
                    file_data = binary_file.read()
    headers = {
                'Content-Type': 'multipart/form-data',
                'x-ms-blob-type' : 'BlockBlob'
            }
    try:
                # Send PUT request
                response = requests.put(files, headers=headers, data=file_data)
                # Check the response
                if response.status_code in [200, 201]:
                    print(f"Successfully uploaded: {csv_file_path}")
                else:
                    print(f"Failed to upload {csv_file_path}. Status code: {response.status_code}, Response: {response.text}")
    except Exception as e:
                print(f"Error uploading {csv_file_path}: {e}")

    fetchDownloadableUrl_path = responsegetPreSignedUrlApi["result"]["filePath"]
    return fetchDownloadableUrl_path


def UpdateUserRole(FilePath,accessToken):
    header_create_user = {
        'X-auth-token': 'bearer ' + accessToken,
        'Content-type':'application/json'
    }
    payload = {
          "file_path": FilePath
        }
    
    url_create_user_api = config.get(environment, 'mentoringthost') + config.get(environment, 'userBulkRoleUpdate')
    response = requests.post(
                url=url_create_user_api,
                headers=header_create_user,
                data=json.dumps(payload)
                )
    if response.status_code == 201:
        print(f"Session creation succeeded for title: {FilePath}")
    else:
        print(f"Session creation failed for title: {FilePath}")
        print(f"Response: {response.text}")



def convert_sheets_to_csv(programFile):
    # Define the output folder
    output_folder = "MentoringDataCSV"
    
    # Create the folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Load the Excel file
    try:
        excel_data = pd.ExcelFile(programFile)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return
    
    # Iterate over each sheet in the Excel file
    for sheet_name in excel_data.sheet_names:
        try:
            # Read the sheet into a DataFrame
            sheet_data = excel_data.parse(sheet_name)
            
            # Define the CSV file path
            csv_file_path = os.path.join(output_folder, f"{sheet_name}.csv")
            
            # Save the DataFrame to CSV
            sheet_data.to_csv(csv_file_path, index=False)
            print(f"Saved {sheet_name} to {csv_file_path}")
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {e}")

def valid_file(param):
    base, ext = os.path.splitext(param)
    if ext.lower() not in ('.xlsx'):
        raise argparse.ArgumentTypeError('File must have a csv extension')
    return param

# function to check environment 
def envCheck():
    try:
        config.get(environment, 'userlogin')
        return True
    except Exception as e:
        print(e)
        return False


def deleteMentoringFolder():
    shutil.rmtree("MentoringDataCSV") 

def updateEnvKeyToFalse():
    env_file = ".env"
    if not os.path.exists(env_file):
        print(".env file not found!")
        return

    with open(env_file, "r") as file:
        lines = file.readlines()

    updated_lines = []
    for line in lines:
        if line.startswith("ENABLE_EMAIL_OTP_VERIFICATION="):
            key, value = line.strip().split("=", 1)
            if value.lower() == "true":
                updated_lines.append(f"{key}=false\n")
            else:
                updated_lines.append(line)
        else:
            updated_lines.append(line)

    updated_session = []
    for line in lines:
        if line.startswith("SESSION_CREATION_MENTOR_LIMIT="):
            key,value = line.strip().split("=",1)
            if value.lower() == "1":
                updated_session.append(f"{key}=5\n")
            else:
                updated_session.append(line)
        else:
            updated_session.append(line)
    print(updated_session,"line no ")
    with open(env_file, "w") as file:
        file.writelines(updated_session)
    with open(env_file, "w") as file:
        file.writelines(updated_lines)

    print(".env file updated successfully.")


def updateEnvKeyToTrue():
    env_file = ".env"
    if not os.path.exists(env_file):
        print(".env file not found!")
        return

    with open(env_file, "r") as file:
        lines = file.readlines()

    updated_lines = []
    for line in lines:
        if line.startswith("ENABLE_EMAIL_OTP_VERIFICATION="):
            key, value = line.strip().split("=", 1)
            if value.lower() == "false":
                updated_lines.append(f"{key}=true\n")
            else:
                updated_lines.append(line)
        else:
            updated_lines.append(line)

    for line in lines:
        if line.startswith("SESSION_CREATION_MENTOR_LIMIT="):
            key, value = line.strip().split("=", 1)
            if value.lower() == "5":
                updated_lines.append(f"{key}=1\n")
            else:
                updated_lines.append(line)
        else:
            updated_lines.append(line)
    
    with open(env_file, "w") as file:
        file.writelines(updated_lines)

    print(".env file updated successfully.")


def updateEnvKeyForSessionsTo5():
    env_file = ".env"
    if not os.path.exists(env_file):
        print(".env file not found!")
        return

    with open(env_file, "r") as file:
        lines = file.readlines()
    updated_session = []
    for line in lines:
        if line.startswith("SESSION_CREATION_MENTOR_LIMIT="):
            key,value = line.strip().split("=",1)
            if value.lower() == "1":
                updated_session.append(f"{key}=5\n")
            else:
                updated_session.append(line)
        else:
            updated_session.append(line)
    with open(env_file, "w") as file:
        file.writelines(updated_session)

    print(".env file updated For Session successfully.")


def updateEnvKeyForSessionsTo1():
    env_file = ".env"
    if not os.path.exists(env_file):
        print(".env file not found!")
        return

    with open(env_file, "r") as file:
        lines = file.readlines()

    updated_lines = []
    for line in lines:
        if line.startswith("SESSION_CREATION_MENTOR_LIMIT="):
            key, value = line.strip().split("=", 1)
            if value.lower() == "5":
                updated_lines.append(f"{key}=1\n")
            else:
                updated_lines.append(line)
        else:
            updated_lines.append(line)
    
    with open(env_file, "w") as file:
        file.writelines(updated_lines)

    print(".env file updated For Session successfully.")

# Main function were all the function def are called
def mainFunc(UserFile):
        # convert_sheets_to_csv(programFile)
        # updateEnvKeyToFalse()
        # updateEnvKeyForSessionsTo5()
        # AdminCreate(UserFile)
        # UserCreate(UserFile)
        accessToken = generateAccessToken()
        # DownloadableUrl_path = getPreSignedUrl(accessToken)
        # UpdateUserRole(DownloadableUrl_path,accessToken)
        # createOrginzation(UserFile,accessToken)
        userRole = sessionTokenGeneration(programFile)
        # setOrgPolicies(UserFile,userRole)
        sessionsFlow(UserFile,userRole)
        CreatingReport(UserFile,userRole)
        deleteMentoringFolder()
        updateEnvKeyToTrue()
        updateEnvKeyForSessionsTo1()

#main execution
start_time = time.time()
parser = argparse.ArgumentParser()
parser.add_argument('--MentoringFile', type=valid_file)
parser.add_argument('--env', '--env')
argument = parser.parse_args()
programFile = argument.MentoringFile
environment = argument.env
millisecond = int(time.time() * 1000)

if envCheck():
    print("=================== Environment set to " + str(environment) + "=====================")
else:
    terminatingMessage(str(environment) + " is an invalid environment")

wbPgm = xlrd.open_workbook(programFile, on_demand=True)
sheetNames = wbPgm.sheet_names()

pgmSheets = ["Organization", "Admin Role", "User Role", "Sessions", "Policies","Update Roles","Sessions Report"]

if len(sheetNames) == len(pgmSheets) and sheetNames == pgmSheets:
    millisecond = int(time.time() * 1000)
    main = mainFunc(programFile)
    end_time = time.time()


print("Execution time in sec : " + str(end_time - start_time))